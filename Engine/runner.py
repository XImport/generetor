import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
# from openpyxl.styles.borders import Border, Side
from Engine.products import PdtTypes
from Engine.transport import (
    ChiffreAffaireTransport,
    Transport,
    CoutTransport,
    MargeTransport,
)
from Engine.chiffredaffaire import ChiffreAffaireNet, chiffreDaffaireCount
from Engine.clients import ClientwithM3, Clientwithtonne, Clients
from Engine.pourcentage import PourcentageFacturation, PourcentageNonFacturation
from Engine.products import TypesFilter, PdtTypes
from Engine.date import DateHandler
from Engine.chargetime import ChargeTime


class Main:
    def __init__(self, file):
        try:

            root_folder = os.path.dirname(os.path.abspath(__file__))
            excel = pd.read_excel(file, sheet_name=sys.argv[2])
            db = pd.read_excel(
                root_folder + "/db/db.xlsx", sheet_name="TypeProducts", index_col=0
            )
            db_pdt = pd.read_excel(
                root_folder + "/db/db.xlsx", sheet_name="PDTS", index_col=0
            )
            db_clts = pd.read_excel(
                root_folder + "/db/db.xlsx", sheet_name="DB", index_col=0
            )
            db_transport = pd.read_excel(
                root_folder + "/db/db.xlsx", sheet_name="TRANSPORT", index_col=0
            )

            pourcentage_fac = pd.read_excel(
                root_folder + "/db/db.xlsx", sheet_name="% FACTURATION", index_col=0
            )

            #############################################################################################################
            dates = DateHandler(excel)
            ticket = excel["N° Ticket"]
            chargeTime = ChargeTime(excel)
            # chargeTime = excel.columns[3]
            bonneLiv = excel["Bon de Livraison"]
            matrecule = excel["Véhicule"]
            bonneCom = excel["Bon de commande"]
            clients = excel["Client"]
            globalclients = Clients(bonneCom, clients)
            types = TypesFilter(excel, db["GRAVES"], db["NOBLES"], db["STERIL"])

            pdts = PdtTypes(excel, db_pdt)
            produits = excel["Produit"]
            net = excel["Net"]
            Qtem3 = excel["Quantité en M3"]
            Qtet =excel[excel.columns[14]]
            # print(Qtet)

            prix_with_m3 = ClientwithM3(globalclients, db_clts, produits.to_list())
            prix_with_tonne = Clientwithtonne(
                globalclients, db_clts, produits.to_list()
            )
            CA = chiffreDaffaireCount(
                net, prix_with_tonne, globalclients, db_clts, Qtem3, prix_with_m3
            )

            prix_transport = Transport(excel, db_transport)
            cout_transport = CoutTransport(
                excel, db_transport, net, globalclients, db_clts, Qtem3, prix_transport
            )

            Catransprt = ChiffreAffaireTransport(
                net, globalclients, db_clts, Qtem3, prix_transport
            )
            CaNet = ChiffreAffaireNet(CA, Catransprt)
            CaNetFac = PourcentageFacturation(globalclients, CaNet, pourcentage_fac)
            CaNonFact = PourcentageNonFacturation(CaNet, CaNetFac)
            Transporteur = excel["Transporteur"]
            Marge = MargeTransport(Catransprt, cout_transport)
            Chantier = excel["Destination"]

            load_data = {
                "Date": dates,
                "Ticket": ticket,
                "Durée de charge": chargeTime,
                "BL": bonneLiv,
                "Matricule": matrecule,
                "BC": bonneCom,
                "CLIENTS": globalclients,
                "Type": types,
                "Pdt": pdts,
                "Produit": produits,
                "Qté en T": net,
                "Qté en m3": Qtem3,
                "Densité": Qtet,
                "Prix en T": prix_with_tonne,
                "Prix en m3": prix_with_m3,
                "CA BRUT": CA,
                "PRIX TRANSPORT": prix_transport,
                "CA TRANSPORT": Catransprt,
                "CA Net": CaNet,
                "CA NET FACT": CaNetFac,
                "CA NON FACT": CaNonFact,
                "COUT TRANSPORT": cout_transport,
                "Marge sur Transport": Marge,
                "Transporteur": Transporteur,
                "Chantier": Chantier,
            }

            data = pd.DataFrame(
                dict([(k, pd.Series(v)) for k, v in load_data.items()])
            )  # B (Better)

            data.to_excel(
                r"C:\Users\ADMIN\Documents\Automation\GeneratedOutput.xlsx", index=False
            )
            wb = openpyxl.load_workbook(
                r"C:\Users\ADMIN\Documents\Automation\GeneratedOutput.xlsx"
            )
            ws = wb["Sheet1"]
            for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                    cell.fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )

            wb.save(r"C:\Users\ADMIN\Documents\Automation\GeneratedOutput.xlsx")

          
        except ValueError:
            print("Day is not existed ...")


if __name__ == "__main__":
    root = Main(sys.argv[1])
